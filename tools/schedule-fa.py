"""
Schedule FA Table A3 Generator
================================
Parses an IBKR annual statement (Flex Query CSV or Activity Statement PDF)
and outputs a ready-to-file Schedule FA Table A3 for Indian Income Tax Return.

All values converted to INR using SBI Telegraphic Transfer Buying Rate (TTBR)
on the relevant transaction date — the source mandated by Indian tax law.

Usage:
    python schedule-fa.py <input.csv|input.pdf> [--year 2025]

Output:
    schedule-fa-table-a3-<year>.csv
    schedule-fa-table-a3-<year>.xlsx  (3 sheets: Table A3, Exchange Rates, Summary)

IBKR Flex Query Setup (for CSV mode, more reliable than PDF):
    Reports → Activity → Flex Queries → Create Query
    Sections: Trades, Open Positions, Dividends
    Format: CSV, Date format: YYYYMMDD
    Trades fields: Symbol, Date/Time, Quantity, TradePrice, Proceeds, Buy/Sell, Asset Category
    Open Positions fields: Symbol, Quantity, CostBasisPrice, MarkPrice, PositionValue
    Dividends fields: Symbol, Date, Amount, Description
"""

import sys
import re
import csv
import json
import argparse
import requests
import numpy as np
import pandas as pd
from io import StringIO
from datetime import datetime, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

TOOLS_DIR = Path(__file__).resolve().parent
TTBR_CACHE = TOOLS_DIR / "usdinr-cache.csv"
GITHUB_API  = "https://api.github.com/repos/prakhar3949/sbi-tt-rates-historical/contents"
GITHUB_RAW  = "https://raw.githubusercontent.com/prakhar3949/sbi-tt-rates-historical/master"
YF_HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

# ── Exchange Rate ─────────────────────────────────────────────────────────────

def _parse_sbi_pdf_rate(pdf_bytes: bytes) -> float | None:
    """
    Parse an SBI FOREX card rate PDF and return the USD TT Buying rate.
    The PDF table has rows like: Currency | TT Buy | TT Sell | ...
    """
    try:
        import pdfplumber, io
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    for row in table:
                        if not row:
                            continue
                        row_str = " ".join(str(c or "").upper() for c in row)
                        if "US DOLLAR" in row_str or "USD" in row_str:
                            # Columns are typically: Currency, TT Buy, TT Sell, Bills Buy, ...
                            # Pick the first numeric value in the reasonable USD/INR range
                            for cell in row[1:]:  # skip the currency name cell
                                val = to_num(str(cell or ""))
                                if 70.0 < val < 120.0:
                                    return val
    except Exception:
        pass
    return None


def _fetch_month_listing(year: str, month: str) -> dict[str, str]:
    """
    Fetch GitHub directory listing for {year}/{month}.
    Returns {YYYYMMDD: raw_pdf_url} for the first PDF found per date.
    File names look like: 2025-06-01-14:15.pdf
    """
    url = f"{GITHUB_API}/{year}/{month}"
    try:
        resp = requests.get(url, timeout=15)
        if resp.status_code != 200:
            return {}
        date_urls: dict[str, str] = {}
        for item in resp.json():
            name = item.get("name", "")
            if not name.endswith(".pdf"):
                continue
            # "2025-06-01-14:15.pdf" → parts[0..2] = year, month, day
            parts = name.replace(".pdf", "").split("-")
            if len(parts) >= 3:
                date_key = f"{parts[0]}{parts[1].zfill(2)}{parts[2].zfill(2)}"
                if date_key not in date_urls:          # keep only the first PDF per day
                    date_urls[date_key] = item["download_url"]
        return date_urls
    except Exception:
        return {}


def load_ttbr(needed_dates: set | None = None) -> dict:
    """
    Load SBI TT Buying rates for USD/INR from the user's GitHub repo:
    https://github.com/prakhar3949/sbi-tt-rates-historical

    Structure: /{year}/{month_number}/{YYYY-MM-DD-HH:MM.pdf}
    Two PDFs per date (morning + afternoon) — we take the first one.

    Args:
        needed_dates: set of YYYYMMDD strings to fetch. If None, returns cache only.
    """
    # Load existing cache
    cached: dict[str, float] = {}
    if TTBR_CACHE.exists():
        df_c = pd.read_csv(TTBR_CACHE)
        cached = dict(zip(df_c["Date"].astype(str), df_c["USDINR"].astype(float)))

    if not needed_dates:
        print(f"  Loaded {len(cached)} cached USD/INR rates")
        return cached

    # Dates we still need
    missing = {d for d in needed_dates if d not in cached}
    if not missing:
        print(f"  All {len(needed_dates)} dates found in cache")
        return cached

    # Group missing dates by year/month
    from collections import defaultdict
    by_month: dict[tuple, list] = defaultdict(list)
    for d in missing:
        by_month[(d[:4], d[4:6])].append(d)

    print(f"  Fetching SBI TT rates for {len(missing)} dates across {len(by_month)} months...")

    # Fetch directory listing for each month then download needed PDFs in parallel
    date_pdf_urls: dict[str, str] = {}
    for (year, month), dates in by_month.items():
        listing = _fetch_month_listing(year, month)
        for d in dates:
            if d in listing:
                date_pdf_urls[d] = listing[d]
            else:
                print(f"    No SBI PDF found for {d} (weekend/holiday — will use nearest prior day)")

    def fetch_and_parse(date_url: tuple[str, str]) -> tuple[str, float | None]:
        date, url = date_url
        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            rate = _parse_sbi_pdf_rate(resp.content)
            return date, rate
        except Exception as e:
            return date, None

    with ThreadPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(fetch_and_parse, (d, u)): d for d, u in date_pdf_urls.items()}
        for fut in as_completed(futures):
            date, rate = fut.result()
            if rate is not None:
                cached[date] = rate
                print(f"    {date}: INR {rate:.2f}")
            else:
                print(f"    {date}: could not parse rate from PDF")

    # Persist updated cache
    rows = [{"Date": k, "USDINR": v} for k, v in sorted(cached.items())]
    pd.DataFrame(rows).to_csv(TTBR_CACHE, index=False)
    print(f"  {len(cached)} rates cached to {TTBR_CACHE}")
    return cached


def get_rate(date_str: str, ttbr: dict, verbose_missing: bool = True) -> float:
    """Return SBI TTBR rate for a given date (YYYYMMDD), walking back up to 7 days."""
    d = datetime.strptime(date_str, "%Y%m%d")
    for offset in range(8):
        key = (d - timedelta(days=offset)).strftime("%Y%m%d")
        if key in ttbr:
            return ttbr[key]
    if verbose_missing:
        print(f"  WARNING: No TTBR rate found near {date_str}, using 84.0 as fallback")
    return 84.0  # approximate fallback if data gap


# ── Yahoo Finance ─────────────────────────────────────────────────────────────

def fetch_ohlcv(ticker: str, period: str = "1y") -> pd.DataFrame | None:
    """Fetch daily OHLCV from Yahoo Chart API."""
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?range={period}&interval=1d"
    try:
        resp = requests.get(url, headers=YF_HEADERS, timeout=15)
        if resp.status_code != 200:
            return None
        data = resp.json()
        result = data.get("chart", {}).get("result", [])
        if not result:
            return None
        r = result[0]
        ts = r["timestamp"]
        q = r["indicators"]["quote"][0]
        df = pd.DataFrame({
            "date": pd.to_datetime(ts, unit="s").strftime("%Y%m%d"),
            "open": q.get("open", []),
            "high": q.get("high", []),
            "low": q.get("low", []),
            "close": q.get("close", []),
            "volume": q.get("volume", []),
        })
        df = df.dropna(subset=["close"])
        return df
    except Exception:
        return None


def fetch_company_info(ticker: str) -> dict:
    """Fetch company name and address from Yahoo quoteSummary."""
    url = f"https://query1.finance.yahoo.com/v10/finance/quoteSummary/{ticker}?modules=assetProfile"
    try:
        resp = requests.get(url, headers=YF_HEADERS, timeout=15)
        if resp.status_code != 200:
            return {}
        data = resp.json()
        profile = data.get("quoteSummary", {}).get("result", [{}])[0].get("assetProfile", {})
        return {
            "name": profile.get("longName") or profile.get("shortName") or ticker,
            "address": ", ".join(filter(None, [
                profile.get("address1", ""),
                profile.get("city", ""),
                profile.get("state", ""),
            ])),
            "zip": profile.get("zip", ""),
            "country": profile.get("country", "United States"),
        }
    except Exception:
        return {}


# ── IBKR Parsers ─────────────────────────────────────────────────────────────

def parse_flex_csv(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Parse IBKR Flex Query CSV.
    Returns (trades_df, positions_df, dividends_df).
    """
    sections: dict[str, list[list[str]]] = {}
    current_section = None
    current_header = None

    with open(path, encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row:
                continue
            section_name = row[0].strip()
            if len(row) < 2:
                continue
            marker = row[1].strip().lower()

            if marker == "header":
                current_section = section_name
                current_header = [c.strip() for c in row[2:]]
                sections.setdefault(current_section, [])
                sections[current_section + "_HEADER"] = current_header
            elif marker == "data" and current_section:
                sections.setdefault(current_section, []).append(row[2:])

    def to_df(name: str) -> pd.DataFrame:
        header = sections.get(name + "_HEADER", [])
        rows = sections.get(name, [])
        if not header or not rows:
            return pd.DataFrame()
        # Pad/trim rows to match header length
        padded = [r[:len(header)] + [""] * max(0, len(header) - len(r)) for r in rows]
        return pd.DataFrame(padded, columns=header)

    trades_df = to_df("Trades")
    positions_df = to_df("Open Positions")
    dividends_df = to_df("Dividends")

    # Normalize trades
    if not trades_df.empty:
        # Filter to stocks only
        if "Asset Category" in trades_df.columns:
            trades_df = trades_df[trades_df["Asset Category"].str.strip().str.lower() == "stocks"]
        # Rename columns to canonical names
        col_map = {}
        for c in trades_df.columns:
            lc = c.lower().replace("/", " ").strip()
            if "symbol" in lc:
                col_map[c] = "Symbol"
            elif "date" in lc and "time" in lc:
                col_map[c] = "DateTime"
            elif lc in ("date",):
                col_map[c] = "DateTime"
            elif "qty" in lc or "quantity" in lc:
                col_map[c] = "Quantity"
            elif "trade" in lc and "price" in lc:
                col_map[c] = "Price"
            elif "proceeds" in lc:
                col_map[c] = "Proceeds"
            elif "buy" in lc and "sell" in lc:
                col_map[c] = "Action"
        trades_df = trades_df.rename(columns=col_map)

    return trades_df, positions_df, dividends_df


def parse_pdf(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Parse IBKR Annual Activity Statement PDF.

    Actual PDF structure observed (pdfplumber):
    - Each section starts a table whose row[0][0] == section name
    - Open Positions: row[1]=column headers, row[2]='Stocks', row[3]='USD', row[4+]=data
    - Trades: first page has row[1]=headers, row[2]='Stocks', row[3]='USD', row[4+]=data;
              continuation pages have row[0]='Trades', row[1+]=data (no repeated headers)
    - Dividends: row[1]=['Date','Description','Amount'], row[2]='USD', row[3+]=data
    """
    try:
        import pdfplumber
    except ImportError:
        print("ERROR: pdfplumber not installed. Run: pip install pdfplumber")
        sys.exit(1)

    SKIP_FIRST_COL = {"", "stocks", "usd", "total", "subtotal"}

    trades_rows: list[list[str]] = []
    positions_rows: list[list[str]] = []
    dividends_rows: list[list[str]] = []
    instrument_names: dict[str, str] = {}   # {ticker: description}
    instrument_types: dict[str, str] = {}   # {ticker: type e.g. ETF/COMMON/ADR}

    # Fixed column indices for Trades (from the observed PDF header):
    # 0=Symbol, 1=Date/Time, 2='', 3=Quantity, 4=T.Price, 5=C.Price,
    # 6=Proceeds, 7=Comm/Fee, 8=Basis, 9=Realized P/L, 10='', 11=MTM P/L, 12=Code
    TRADE_COLS = {
        "Symbol": 0, "DateTime": 1, "Quantity": 3,
        "Price": 4, "Proceeds": 6,
    }

    # Fixed column indices for Open Positions:
    # 0=Symbol, 1=Quantity, 2=Mult, 3=Cost Price, 4=Cost Basis, 5=Close Price, 6=Value
    POS_COLS = {
        "Symbol": 0, "Quantity": 1, "CostBasisPrice": 3, "MarkPrice": 5,
    }

    # Dividends: 0=Date, 1=Description, 2=Amount
    DIV_COLS = {"Date": 0, "Description": 1, "Amount": 2}

    # Financial Instrument Information: 0=Symbol, 1=Description, ...
    FIN_INST_SECTION = "Financial Instrument Information"

    print("  Parsing PDF pages...")
    with pdfplumber.open(path) as pdf:
        print(f"  Total pages: {len(pdf.pages)}")
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table or not table[0] or not table[0][0]:
                    continue
                section = str(table[0][0]).strip()

                # ── Open Positions ──
                if section == "Open Positions":
                    # data starts at row index 4 (after header, Stocks, USD rows)
                    for row in table[4:]:
                        sym = str(row[0] or "").strip()
                        if not sym or sym.lower() in SKIP_FIRST_COL or sym.lower().startswith("total"):
                            continue
                        positions_rows.append([str(c or "").strip() for c in row])

                # ── Trades ──
                elif section == "Trades":
                    if len(table) < 2:
                        continue
                    # Determine if row[1] is a column-header row or a data row
                    row1_first = str(table[1][0] or "").strip()
                    if row1_first == "Symbol":
                        # First occurrence of this section on this page: skip header+Stocks+USD
                        data_start = 4
                    else:
                        # Continuation page: data starts right after the section-name row
                        data_start = 1
                    for row in table[data_start:]:
                        if len(row) <= TRADE_COLS["Quantity"]:
                            continue
                        sym = str(row[TRADE_COLS["Symbol"]] or "").strip()
                        if not sym or sym.lower() in SKIP_FIRST_COL or sym.lower().startswith("total"):
                            continue
                        trades_rows.append([str(c or "").strip() for c in row])

                # ── Dividends ──
                elif section == "Dividends":
                    # row[1] = ['Date', 'Description', 'Amount'], row[2] = 'USD', row[3+] = data
                    for row in table[3:]:
                        if not row or not row[0]:
                            continue
                        date_val = str(row[0] or "").strip()
                        if not date_val or date_val.lower() in ("usd", "date", ""):
                            continue
                        dividends_rows.append([str(c or "").strip() for c in row])

                # ── Financial Instrument Information ──
                # Columns: 0=Symbol, 1=Description, 2=Conid, 3=SecurityID,
                #          4=Underlying, 5=ListingExch, 6=Multiplier, 7=Type, 8=Code
                # First page:        row[0]=header, row[1]=col names, row[2]='Stocks', row[3+]=data
                # Continuation pages: row[0]=header, row[1]=col names, row[2+]=data
                elif section == FIN_INST_SECTION:
                    row2_first = str(table[2][0] or "").strip().lower() if len(table) > 2 else ""
                    data_start = 3 if row2_first == "stocks" else 2
                    for row in table[data_start:]:
                        if not row or len(row) < 2:
                            continue
                        sym = str(row[0] or "").strip()
                        desc = str(row[1] or "").strip()
                        itype = str(row[7] or "").strip() if len(row) > 7 else ""
                        if not sym or sym.lower() in SKIP_FIRST_COL or sym.lower().startswith("total"):
                            continue
                        if sym not in instrument_names:
                            instrument_names[sym] = desc.replace("\n", " ").strip()
                            instrument_types[sym] = itype

    # Build DataFrames using fixed column positions
    def rows_to_df(rows: list[list[str]], col_map: dict[str, int]) -> pd.DataFrame:
        if not rows:
            return pd.DataFrame()
        records = []
        for row in rows:
            rec = {}
            for col_name, idx in col_map.items():
                rec[col_name] = row[idx] if idx < len(row) else ""
            records.append(rec)
        return pd.DataFrame(records)

    trades_df = rows_to_df(trades_rows, TRADE_COLS)
    positions_df = rows_to_df(positions_rows, POS_COLS)
    dividends_df = rows_to_df(dividends_rows, DIV_COLS)

    print(f"  Parsed: {len(trades_rows)} trade rows, {len(positions_rows)} positions, {len(dividends_rows)} dividends, {len(instrument_names)} instrument names")


    # Add Action column to trades (infer from Quantity sign: negative = SELL)
    if not trades_df.empty and "Quantity" in trades_df.columns:
        def infer_action(qty_str):
            try:
                return "SELL" if float(qty_str.replace(",", "")) < 0 else "BUY"
            except ValueError:
                return ""
        trades_df["Action"] = trades_df["Quantity"].apply(infer_action)
        # Make Quantity always positive (sign now in Action)
        trades_df["Quantity"] = trades_df["Quantity"].apply(
            lambda x: x.lstrip("-").strip()
        )

    return trades_df, positions_df, dividends_df, instrument_names, instrument_types


def parse_input(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict, dict]:
    ext = path.suffix.lower()
    if ext == ".pdf":
        print("Input: PDF (Annual Activity Statement)")
        return parse_pdf(path)
    elif ext in (".csv", ".txt"):
        print("Input: CSV (Flex Query export)")
        trades_df, positions_df, dividends_df = parse_flex_csv(path)
        return trades_df, positions_df, dividends_df, {}, {}
    else:
        print(f"ERROR: Unsupported file type '{ext}'. Use .csv or .pdf")
        sys.exit(1)


# ── Numeric helpers ───────────────────────────────────────────────────────────

def to_num(v) -> float:
    """Convert string/numeric to float, stripping commas and currency symbols."""
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return 0.0


def parse_date(v: str) -> str | None:
    """Parse various date formats into YYYYMMDD. Returns None on failure."""
    # Normalize: strip, replace newline, take first token before space/semicolon/comma
    v = str(v).strip().replace("\n", " ").replace(",", " ")
    v = v.split(";")[0].split(" ")[0].split("T")[0].strip()
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(v, fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    return None


# ── Core Processing ───────────────────────────────────────────────────────────

def build_trades(trades_df: pd.DataFrame, year: int) -> pd.DataFrame:
    """
    Normalize trade records into a canonical DataFrame with columns:
    Symbol, Date, Quantity (signed: + buy, - sell), Price, Proceeds, Action
    """
    if trades_df.empty:
        return pd.DataFrame(columns=["Symbol", "Date", "Quantity", "Price", "Proceeds", "Action"])

    rows = []
    for _, row in trades_df.iterrows():
        sym = str(row.get("Symbol", "")).strip().replace(" ", "").upper()
        if not sym or sym.startswith("TOTAL"):
            continue

        date_raw = str(row.get("DateTime", row.get("Date", ""))).strip()
        date = parse_date(date_raw)
        if not date:
            continue

        # Filter to the relevant year
        if not date.startswith(str(year)):
            continue

        qty = to_num(row.get("Quantity", 0))
        price = to_num(row.get("Price", row.get("TradePrice", 0)))
        proceeds = to_num(row.get("Proceeds", 0))
        action = str(row.get("Action", row.get("Buy/Sell", ""))).strip().upper()

        # Infer action from quantity sign if not present
        if not action:
            action = "BUY" if qty > 0 else "SELL"
        elif action in ("B", "BOT", "BUY"):
            action = "BUY"
        elif action in ("S", "SLD", "SELL"):
            action = "SELL"

        rows.append({
            "Symbol": sym,
            "Date": date,
            "Quantity": abs(qty),
            "Price": price,
            "Proceeds": abs(proceeds),
            "Action": action,
        })

    cols = ["Symbol", "Date", "Quantity", "Price", "Proceeds", "Action"]
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)


def build_positions(positions_df: pd.DataFrame) -> dict:
    """
    Parse Dec 31 open positions.
    Returns {ticker: {"shares": float, "cost_basis": float, "close_price": float}}
    """
    result = {}
    if positions_df.empty:
        return result

    for _, row in positions_df.iterrows():
        sym = str(row.get("Symbol", "")).strip().upper()
        if not sym or sym in ("SYMBOL", "TOTAL"):
            continue

        # Try various column name spellings
        shares = to_num(row.get("Quantity", row.get("Qty", row.get("Position", 0))))
        cost = to_num(row.get("CostBasisPrice", row.get("Cost Price", row.get("Avg Cost", 0))))
        close = to_num(row.get("MarkPrice", row.get("Close Price", row.get("Mark Price", 0))))

        if shares != 0:
            result[sym] = {"shares": shares, "cost_basis": cost, "close_price": close}

    return result


def build_dividends(dividends_df: pd.DataFrame, year: int) -> dict:
    """
    Parse dividends.
    Returns {ticker: [(date, amount), ...]}
    """
    result: dict[str, list] = {}
    if dividends_df.empty:
        return result

    for _, row in dividends_df.iterrows():
        desc = str(row.get("Description", row.get("Symbol", ""))).strip()
        # Extract ticker from description like "AMZN (US67066G1040) Cash Dividend..."
        sym_match = re.match(r"^([A-Z]+)", desc.upper())
        sym = sym_match.group(1) if sym_match else str(row.get("Symbol", "")).strip().upper()
        if not sym:
            continue

        date_raw = str(row.get("Date", row.get("Pay Date", ""))).strip()
        date = parse_date(date_raw)
        if not date or not date.startswith(str(year)):
            continue

        amount = to_num(row.get("Amount", 0))
        if amount <= 0:
            continue

        result.setdefault(sym, []).append((date, amount))

    return result


def build_shares_timeline(ticker: str, trades: pd.DataFrame, initial_shares: float, year: int, ohlcv: pd.DataFrame) -> pd.DataFrame:
    """
    For a ticker, build a day-by-day shares-held series across the year.
    Used to compute peak market value.
    """
    # Start with shares from prior year (initial_shares)
    year_trades = trades[trades["Symbol"] == ticker].copy()
    year_trades = year_trades.sort_values("Date")

    # Merge trades onto OHLCV calendar
    if ohlcv is None or ohlcv.empty:
        return pd.DataFrame()

    df = ohlcv.copy()
    df = df[df["date"].str.startswith(str(year))]

    shares = initial_shares
    trade_by_date: dict[str, float] = {}
    for _, t in year_trades.iterrows():
        delta = t["Quantity"] if t["Action"] == "BUY" else -t["Quantity"]
        trade_by_date[t["Date"]] = trade_by_date.get(t["Date"], 0) + delta

    held_series = []
    for _, row in df.iterrows():
        shares += trade_by_date.get(row["date"], 0)
        held_series.append(max(0, shares))  # can't go negative

    df = df.copy()
    df["shares_held"] = held_series
    return df


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Schedule FA Table A3 from IBKR statement")
    parser.add_argument("input", help="IBKR statement file (.csv or .pdf)")
    parser.add_argument("--year", type=int, default=2025, help="Tax year (default: 2025)")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if not input_path.exists():
        print(f"ERROR: File not found: {input_path}")
        sys.exit(1)

    year = args.year
    print(f"\nSchedule FA Table A3 Generator — {year}")
    print("=" * 50)

    # ── Step 1: Parse input
    trades_df_raw, positions_df_raw, dividends_df_raw, instrument_names, instrument_types = parse_input(input_path)

    trades = build_trades(trades_df_raw, year)
    positions = build_positions(positions_df_raw)  # {ticker: {shares, cost_basis, close_price}}
    dividends = build_dividends(dividends_df_raw, year)

    if instrument_names:
        print(f"  Instrument names from PDF: {len(instrument_names)} tickers")

    def nature_of_entity(ticker: str) -> str:
        return "ETF" if instrument_types.get(ticker) == "ETF" else "Listed Company"

    # ── Step 2: Ticker universe
    all_tickers = sorted(set(trades["Symbol"].unique()) | set(positions.keys()))
    if not all_tickers:
        print("ERROR: No equity positions found. Check input file format.")
        sys.exit(1)

    print(f"\nTickers found: {len(all_tickers)}")
    print(f"  {', '.join(all_tickers)}")

    # ── Step 3: Collect all dates we need exchange rates for
    needed_dates: set[str] = set()
    needed_dates.add(f"{year}1231")   # Dec 31 closing
    needed_dates.add(f"{year}0101")   # Jan 1 (carryover positions)
    for _, row in trades.iterrows():
        if row["Date"]:
            needed_dates.add(row["Date"])
    for divs in dividends.values():
        for d, _ in divs:
            needed_dates.add(d)

    # ── Step 3b: SBI TT rates from GitHub repo
    print(f"\nLoading SBI TT rates for {len(needed_dates)} transaction dates...")
    ttbr = load_ttbr(needed_dates)

    # ── Step 4: Fetch daily prices (parallel)
    print(f"\nFetching 1-year daily prices for {len(all_tickers)} tickers...")
    ohlcv_map: dict[str, pd.DataFrame | None] = {}

    def fetch_ticker(t):
        return t, fetch_ohlcv(t, period=f"{year}y" if year != 2025 else "1y")

    with ThreadPoolExecutor(max_workers=5) as ex:
        futures = {ex.submit(fetch_ticker, t): t for t in all_tickers}
        for fut in as_completed(futures):
            t, df = fut.result()
            ohlcv_map[t] = df
            status = f"{len(df)} days" if df is not None and not df.empty else "NO DATA"
            print(f"  {t}: {status}")

    company_map: dict[str, dict] = {}  # address/zip still fetched per-ticker on demand below

    # ── Step 6: Compute Table A3 rows (one row per buy lot = per ticker+date)
    print("\nComputing Table A3 rows...")

    table_a3_rows = []
    exchange_rates_audit = []
    warnings = []

    dec31_date = f"{year}1231"

    def _inr(val):
        return round(val, 2) if val is not None else "No price data"

    def _usd(val):
        return round(val, 2) if val is not None else "No price data"

    def _rate(val):
        return round(val, 4) if val is not None else "N/A"

    def fifo_allocate(lots: list[dict], sells: pd.DataFrame) -> list[dict]:
        """
        Allocate sells to buy lots using FIFO.
        Each lot dict has: date, qty, price. Returns lots with added keys:
          remaining_qty  — shares from this lot still held at Dec 31
          lot_sells      — list of {date, qty, proceeds_usd} allocated to this lot
        """
        for lot in lots:
            lot["remaining_qty"] = lot["qty"]
            lot["lot_sells"] = []

        for _, sell in sells.sort_values("Date").iterrows():
            sell_qty_left = sell["Quantity"]
            proceeds_per_share = sell["Proceeds"] / sell["Quantity"] if sell["Quantity"] > 0 else 0.0
            for lot in lots:
                if sell_qty_left <= 0:
                    break
                if lot["remaining_qty"] <= 0:
                    continue
                taken = min(lot["remaining_qty"], sell_qty_left)
                lot["remaining_qty"] -= taken
                sell_qty_left -= taken
                lot["lot_sells"].append({
                    "date": sell["Date"],
                    "qty": taken,
                    "proceeds_usd": taken * proceeds_per_share,
                })
        return lots

    def lot_peak(lot_date: str, lot_qty: float, lot_sells: list[dict],
                 ohlcv: pd.DataFrame, year: int) -> tuple:
        """
        Compute peak INR value for one lot.
        Shares held from lot_date → reduce by FIFO sells as they occur.
        Returns (peak_usd, peak_inr, rate_peak, peak_date_display).
        """
        if ohlcv is None or ohlcv.empty:
            return None, None, None, ""

        # Build a sell-drain schedule: {date: qty_sold_from_this_lot}
        drain: dict[str, float] = {}
        for s in lot_sells:
            drain[s["date"]] = drain.get(s["date"], 0) + s["qty"]

        best_inr = 0.0
        peak_usd = peak_inr = rate_peak = None
        peak_date_display = ""
        shares = 0.0  # shares from this lot held on each day

        for _, row in ohlcv.iterrows():
            d = row["date"]
            if d < lot_date:
                continue
            if not d.startswith(str(year)):
                continue
            if d == lot_date:
                shares = lot_qty  # lot arrives on buy date
            shares -= drain.get(d, 0)
            shares = max(0.0, shares)
            if shares <= 0 or row["high"] is None:
                continue
            rate = get_rate(d, ttbr, verbose_missing=False)
            day_usd = shares * float(row["high"])
            val_inr = day_usd * rate
            if val_inr > best_inr:
                best_inr = val_inr
                peak_usd = day_usd
                peak_inr = val_inr
                rate_peak = rate
                peak_date_display = datetime.strptime(d, "%Y%m%d").strftime("%d/%m/%Y")

        return peak_usd, peak_inr, rate_peak, peak_date_display

    for ticker in all_tickers:
        ticker_trades = trades[trades["Symbol"] == ticker].copy()
        ticker_divs = dividends.get(ticker, [])
        pos = positions.get(ticker, {})
        ohlcv = ohlcv_map.get(ticker)
        info = company_map.get(ticker, {})

        raw_buys = ticker_trades[ticker_trades["Action"] == "BUY"].sort_values("Date")
        sells = ticker_trades[ticker_trades["Action"] == "SELL"].sort_values("Date")
        shares_at_dec31 = pos.get("shares", 0)

        # ── Carried-from-prior-year position (no buys this year) → single row
        has_buys_this_year = not raw_buys.empty
        shares_at_year_start = 0.0
        if ticker in positions and not has_buys_this_year:
            year_sells = sells["Quantity"].sum()
            shares_at_year_start = pos.get("shares", 0) + year_sells

        if not has_buys_this_year:
            if ticker not in positions:
                continue  # nothing to report
            # Single carried-forward row
            acq_display = f"Before 01/01/{year}"
            initial_usd = shares_at_year_start * pos.get("cost_basis", 0)
            rate_acq = get_rate(f"{year}0101", ttbr)
            initial_inr = initial_usd * rate_acq
            exchange_rates_audit.append((f"{year}0101", rate_acq, f"{ticker} carried forward"))

            # Peak for full carried position
            p_usd = p_inr = r_peak = None
            peak_disp = ""
            if ohlcv is not None and not ohlcv.empty:
                drain_cf: dict[str, float] = {}
                for _, s in sells.iterrows():
                    drain_cf[s["Date"]] = drain_cf.get(s["Date"], 0) + s["Quantity"]
                sh = shares_at_year_start
                best_inr = 0.0
                for _, row in ohlcv.iterrows():
                    d = row["date"]
                    if not d.startswith(str(year)):
                        continue
                    sh -= drain_cf.get(d, 0)
                    sh = max(0.0, sh)
                    if sh <= 0 or row["high"] is None:
                        continue
                    rate = get_rate(d, ttbr, verbose_missing=False)
                    day_usd = sh * float(row["high"])
                    val_inr = day_usd * rate
                    if val_inr > best_inr:
                        best_inr = val_inr
                        p_usd, p_inr, r_peak = day_usd, val_inr, rate
                        peak_disp = datetime.strptime(d, "%Y%m%d").strftime("%d/%m/%Y")
                        exchange_rates_audit.append((d, rate, f"{ticker} peak day"))
            else:
                warnings.append(f"{ticker}: No price data available — peak value cannot be determined")

            # Closing
            if shares_at_dec31 > 0:
                close_price = pos.get("close_price", 0)
                if close_price == 0 and ohlcv is not None and not ohlcv.empty:
                    lr = ohlcv[ohlcv["date"] <= dec31_date].tail(1)
                    if not lr.empty:
                        close_price = float(lr["close"].values[0])
                r_close = get_rate(dec31_date, ttbr)
                cl_usd = shares_at_dec31 * close_price
                cl_inr = cl_usd * r_close
                cl_date_disp = datetime.strptime(dec31_date, "%Y%m%d").strftime("%d/%m/%Y")
                exchange_rates_audit.append((dec31_date, r_close, f"{ticker} Dec 31 closing"))
            else:
                cl_usd = cl_inr = 0.0
                r_close = None
                cl_date_disp = "N/A"

            # Dividends on the single row
            div_usd = div_inr = 0.0
            for dv_date, dv_amt in ticker_divs:
                r = get_rate(dv_date, ttbr)
                div_inr += dv_amt * r
                div_usd += dv_amt
                exchange_rates_audit.append((dv_date, r, f"{ticker} dividend"))
            r_div = (div_inr / div_usd) if div_usd else None

            # Proceeds on the single row
            proc_usd = proc_inr = 0.0
            for _, sell in sells.iterrows():
                r = get_rate(sell["Date"], ttbr)
                proc_inr += sell["Proceeds"] * r
                proc_usd += sell["Proceeds"]
                exchange_rates_audit.append((sell["Date"], r, f"{ticker} sale"))
            r_sell = (proc_inr / proc_usd) if proc_usd else None

            table_a3_rows.append({
                "Sl. No.": len(table_a3_rows) + 1,
                "Country Name": "United States",
                "Country Code": "US",
                "Name of Entity": instrument_names.get(ticker, "NO NAME"),
                "Address of Entity": info.get("address", ""),
                "Zip Code": info.get("zip", ""),
                "Nature of Entity": nature_of_entity(ticker),
                "Date of Acquisition": acq_display,
                "Initial Value (USD)": _usd(initial_usd),
                "Rate on Acquisition Date": _rate(rate_acq),
                "Initial Value of Investment (INR)": _inr(initial_inr),
                "Peak Value (USD)": _usd(p_usd),
                "Peak Date": peak_disp if peak_disp else "No price data",
                "Rate on Peak Date": _rate(r_peak),
                "Peak Value During Year (INR)": _inr(p_inr),
                "Closing Value (USD)": _usd(cl_usd),
                "Closing Rate Date": cl_date_disp,
                "Rate on Closing Date": _rate(r_close),
                "Closing Value of Investment (INR)": _inr(cl_inr),
                "Total Gross Income/Dividends (USD)": _usd(div_usd) if div_usd else 0.0,
                "Avg Rate on Dividend Dates": _rate(r_div),
                "Total Gross Income (Dividends) (INR)": _inr(div_inr),
                "Total Gross Proceeds (USD)": _usd(proc_usd) if proc_usd else 0.0,
                "Avg Rate on Sale Dates": _rate(r_sell),
                "Total Gross Proceeds from Sale (INR)": _inr(proc_inr),
            })
            continue

        # ── Lots bought this year: group by date (same-day buys = one lot)
        lots_by_date = (
            raw_buys.groupby("Date", sort=True)
            .agg(qty=("Quantity", "sum"), price=("Price", lambda x: (x * raw_buys.loc[x.index, "Quantity"]).sum() / raw_buys.loc[x.index, "Quantity"].sum()))
            .reset_index()
            .rename(columns={"Date": "date"})
            .to_dict("records")
        )

        # FIFO: allocate sells to lots
        lots = fifo_allocate(lots_by_date, sells)

        # Dividends assigned to the first (earliest) lot only
        div_usd_total = div_inr_total = 0.0
        r_div_total = None
        for dv_date, dv_amt in ticker_divs:
            r = get_rate(dv_date, ttbr)
            div_inr_total += dv_amt * r
            div_usd_total += dv_amt
            exchange_rates_audit.append((dv_date, r, f"{ticker} dividend"))
        r_div_total = (div_inr_total / div_usd_total) if div_usd_total else None

        if ohlcv is None or ohlcv.empty:
            warnings.append(f"{ticker}: No price data available — peak value cannot be determined")

        for lot_idx, lot in enumerate(lots):
            lot_date = lot["date"]
            lot_qty = lot["qty"]
            lot_price = lot["price"]
            lot_sells = lot["lot_sells"]
            lot_remaining = lot["remaining_qty"]

            acq_display = datetime.strptime(lot_date, "%Y%m%d").strftime("%d/%m/%Y")

            # Initial value: lot qty × lot price × TTBR on buy date
            buy_rate = get_rate(lot_date, ttbr)
            initial_usd = lot_qty * lot_price
            initial_inr = initial_usd * buy_rate
            exchange_rates_audit.append((lot_date, buy_rate, f"{ticker} lot buy {acq_display}"))

            # Peak value for this lot
            p_usd, p_inr, r_peak, peak_disp = lot_peak(
                lot_date, lot_qty, lot_sells, ohlcv, year
            )
            if p_inr is not None:
                exchange_rates_audit.append((
                    datetime.strptime(peak_disp, "%d/%m/%Y").strftime("%Y%m%d"),
                    r_peak, f"{ticker} lot {acq_display} peak"
                ))

            # Closing value for this lot
            if lot_remaining > 0:
                close_price = pos.get("close_price", 0)
                if close_price == 0 and ohlcv is not None and not ohlcv.empty:
                    lr = ohlcv[ohlcv["date"] <= dec31_date].tail(1)
                    if not lr.empty:
                        close_price = float(lr["close"].values[0])
                r_close = get_rate(dec31_date, ttbr)
                cl_usd = lot_remaining * close_price
                cl_inr = cl_usd * r_close
                cl_date_disp = datetime.strptime(dec31_date, "%Y%m%d").strftime("%d/%m/%Y")
                exchange_rates_audit.append((dec31_date, r_close, f"{ticker} lot {acq_display} Dec 31"))
            else:
                cl_usd = cl_inr = 0.0
                r_close = None
                cl_date_disp = "N/A"

            # Proceeds for this lot (FIFO-allocated sells)
            proc_usd = proc_inr = 0.0
            for ls in lot_sells:
                r = get_rate(ls["date"], ttbr)
                proc_inr += ls["proceeds_usd"] * r
                proc_usd += ls["proceeds_usd"]
                exchange_rates_audit.append((ls["date"], r, f"{ticker} lot {acq_display} sale"))
            r_sell = (proc_inr / proc_usd) if proc_usd else None

            # Dividends: assign to first lot only, zero on subsequent lots
            if lot_idx == 0:
                row_div_usd = div_usd_total
                row_div_inr = div_inr_total
                row_r_div = r_div_total
            else:
                row_div_usd = 0.0
                row_div_inr = 0.0
                row_r_div = None

            table_a3_rows.append({
                "Sl. No.": len(table_a3_rows) + 1,
                "Country Name": "United States",
                "Country Code": "US",
                "Name of Entity": instrument_names.get(ticker, "NO NAME"),
                "Address of Entity": info.get("address", ""),
                "Zip Code": info.get("zip", ""),
                "Nature of Entity": nature_of_entity(ticker),
                "Date of Acquisition": acq_display,
                "Initial Value (USD)": _usd(initial_usd),
                "Rate on Acquisition Date": _rate(buy_rate),
                "Initial Value of Investment (INR)": _inr(initial_inr),
                "Peak Value (USD)": _usd(p_usd),
                "Peak Date": peak_disp if peak_disp else "No price data",
                "Rate on Peak Date": _rate(r_peak),
                "Peak Value During Year (INR)": _inr(p_inr),
                "Closing Value (USD)": _usd(cl_usd),
                "Closing Rate Date": cl_date_disp,
                "Rate on Closing Date": _rate(r_close),
                "Closing Value of Investment (INR)": _inr(cl_inr),
                "Total Gross Income/Dividends (USD)": _usd(row_div_usd) if row_div_usd else 0.0,
                "Avg Rate on Dividend Dates": _rate(row_r_div),
                "Total Gross Income (Dividends) (INR)": _inr(row_div_inr),
                "Total Gross Proceeds (USD)": _usd(proc_usd) if proc_usd else 0.0,
                "Avg Rate on Sale Dates": _rate(r_sell),
                "Total Gross Proceeds from Sale (INR)": _inr(proc_inr),
            })

    if not table_a3_rows:
        print("ERROR: No Table A3 rows computed. Check input file.")
        sys.exit(1)

    # ── Step 7: Totals + Output
    output_stem = input_path.parent / f"schedule-fa-table-a3-{year}"
    csv_path = output_stem.with_suffix(".csv")
    xlsx_path = output_stem.with_suffix(".xlsx")

    a3_df = pd.DataFrame(table_a3_rows)

    total_initial = pd.to_numeric(a3_df["Initial Value of Investment (INR)"], errors="coerce").sum()
    total_peak    = pd.to_numeric(a3_df["Peak Value During Year (INR)"], errors="coerce").sum()
    total_closing = pd.to_numeric(a3_df["Closing Value of Investment (INR)"], errors="coerce").sum()
    total_income  = pd.to_numeric(a3_df["Total Gross Income (Dividends) (INR)"], errors="coerce").sum()
    total_proceeds= pd.to_numeric(a3_df["Total Gross Proceeds from Sale (INR)"], errors="coerce").sum()

    # CSV
    a3_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"\nCSV saved: {csv_path}")

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ─ Sheet 1: Table A3
        ws = wb.active
        ws.title = "Table A3"
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        inr_fmt = '#,##0.00'
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(a3_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = border

        ws.row_dimensions[1].height = 40

        for row_idx, row in enumerate(a3_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                col_name = a3_df.columns[col_idx - 1]
                if "(INR)" in col_name and isinstance(value, (int, float)):
                    cell.number_format = inr_fmt
                    cell.alignment = Alignment(horizontal="right")

        # Auto-fit columns
        for col_idx in range(1, len(a3_df.columns) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, len(table_a3_rows) + 2)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"

        # ─ Sheet 2: Exchange Rates Audit
        ws2 = wb.create_sheet("Exchange Rates Used")
        ws2.append(["Date", "USD/INR TTBR Rate", "Used For"])
        ws2["A1"].font = Font(bold=True)
        ws2["B1"].font = Font(bold=True)
        ws2["C1"].font = Font(bold=True)

        seen_audit = set()
        for date_str, rate, note in sorted(set(
            (d, r, n) for d, r, n in exchange_rates_audit
        )):
            key = (date_str, note)
            if key in seen_audit:
                continue
            seen_audit.add(key)
            display_date = datetime.strptime(date_str, "%Y%m%d").strftime("%d/%m/%Y") if len(date_str) == 8 else date_str
            ws2.append([display_date, rate, note])

        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 20
        ws2.column_dimensions["C"].width = 40

        # ─ Sheet 3: Summary
        ws3 = wb.create_sheet("Summary")
        ws3["A1"] = f"Schedule FA Table A3 Summary — {year}"
        ws3["A1"].font = Font(bold=True, size=13)

        summary_rows = [
            ("", ""),
            ("Metric", "Amount (INR)"),
            ("Number of Foreign Equity Positions", len(table_a3_rows)),
            ("Total Initial Investment Cost", total_initial),
            ("Total Peak Value During Year", total_peak),
            ("Total Closing Value (Dec 31)", total_closing),
            ("Total Dividend Income Received", total_income),
            ("Total Sale Proceeds", total_proceeds),
            ("", ""),
            ("Tax Year (Calendar Year)", year),
            ("Assessment Year", f"{year+1}-{str(year+2)[2:]}"),
            ("Currency Rates Source", "USD/INR Daily Rate — Yahoo Finance (USDINR=X)"),
            ("All values in", "Indian Rupees (INR)"),
        ]

        for i, (label, value) in enumerate(summary_rows, 1):
            ws3.cell(row=i+1, column=1, value=label)
            ws3.cell(row=i+1, column=2, value=value)
            if label == "Metric":
                ws3.cell(row=i+1, column=1).font = Font(bold=True)
                ws3.cell(row=i+1, column=2).font = Font(bold=True)
            elif isinstance(value, float):
                ws3.cell(row=i+1, column=2).number_format = inr_fmt

        ws3.column_dimensions["A"].width = 45
        ws3.column_dimensions["B"].width = 25

        wb.save(xlsx_path)
        print(f"Excel saved: {xlsx_path}")

    except ImportError:
        print("  openpyxl not installed — Excel output skipped. Install with: pip install openpyxl")

    # ── Summary
    print(f"\n{'='*50}")
    print(f"Schedule FA Table A3 - {year} Summary")
    print(f"{'='*50}")
    print(f"Positions: {len(table_a3_rows)}")
    print(f"Total Initial Cost:  INR{total_initial:>14,.0f}")
    print(f"Total Peak Value:    INR{total_peak:>14,.0f}")
    print(f"Total Closing Value: INR{total_closing:>14,.0f}")
    print(f"Total Dividends:     INR{total_income:>14,.0f}")
    print(f"Total Sale Proceeds: INR{total_proceeds:>14,.0f}")

    if warnings:
        print(f"\nWarnings ({len(warnings)}):")
        for w in warnings:
            print(f"  WARN  {w}")

    print(f"\nFiles written:")
    print(f"  {csv_path}")
    print(f"  {xlsx_path}")
    print(f"\nAssessment Year: AY {year+1}-{str(year+2)[2:]}  (IBKR calendar year {year})")


if __name__ == "__main__":
    main()

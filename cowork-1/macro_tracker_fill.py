"""
macro_tracker_fill.py — populate macro_thesis_tracker.xlsx with macro economic data.

Companion to macro_tracker.py (which builds the empty workbook). This script:
  1. Opens the existing .xlsx
  2. Fetches FRED + Yahoo data for auto-fetchable metrics
  3. Merges in manual_overrides.json for series with no programmatic feed
  4. Writes values into the correct cells (preserving the blue-input font)
  5. Prints a summary: filled / manual-required / missing

Re-runnable monthly. Idempotent — each run overwrites prior values, which is correct
because FRED revises historical data and yfinance refreshes with each new earnings print.

Usage: python macro_tracker_fill.py [--workbook macro_thesis_tracker.xlsx]
"""

from __future__ import annotations

import argparse
import io
import json
import sys
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font

HERE = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = HERE / "macro_thesis_tracker.xlsx"
OVERRIDES_PATH = HERE / "manual_overrides.json"

UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
BLUE_INPUT_FONT = Font(name="Arial", size=10, color="0000FF")

# ─── Periods (must match headers in macro_tracker.py) ───
MONTHS = ["Jan-25", "Feb-25", "Mar-25", "Apr-25", "May-25", "Jun-25", "Jul-25", "Aug-25",
          "Sep-25", "Oct-25", "Nov-25", "Dec-25",
          "Jan-26", "Feb-26", "Mar-26", "Apr-26", "May-26"]

QUARTERS = ["Q1-24", "Q2-24", "Q3-24", "Q4-24", "Q1-25", "Q2-25", "Q3-25", "Q4-25",
            "Q1-26", "Q2-26"]


# ─── Metric map — strategy + parameter for each label ───
# Strategies:
#   FRED_YOY            — fetch FRED series, YoY % for the period
#   FRED_YOY_DIFF       — (s1, s2) tuple: YoY(s1) - YoY(s2). Used for real wage
#                          growth (nominal AHE YoY minus CPI YoY).
#   FRED_MONTHLY_AVG    — fetch FRED daily series, monthly mean
#   FRED_LATEST         — series already at period freq, return value
#   FRED_QTRLY_AVG      — quarterly mean of daily/monthly series
#   FRED_QTRLY_YOY      — quarterly YoY %
#   FRED_QTRLY_YOY_DIFF — (s1, s2) tuple: YoY(s1) - YoY(s2). Used for the
#                          ULC-minus-Productivity disinflationary-offset signal.
#   FRED_QTRLY_LATEST   — quarterly series, return value
#   YAHOO_CAPEX_QTRLY   — quarterly capex in $B from yfinance cashflow
#   YAHOO_CAPEX_PCT     — capex / revenue % from yfinance
#   YAHOO_CAPEX_COMBINED— sum of capex across multiple tickers
#   DERIVED_QOQ         — QoQ % derived from another metric in this workbook
#   MANUAL              — handled via manual_overrides.json

MONTHLY_METRIC_MAP: dict[str, tuple[str, object]] = {
    # Inflation
    "Core PCE (YoY)":                                      ("FRED_YOY", "PCEPILFE"),
    "Core CPI (YoY)":                                      ("FRED_YOY", "CPILFESL"),
    # Supercore = Services less energy services (closest standard BLS series; Powell's
    # exact "supercore" derivation also strips shelter, which we approximate here)
    "Supercore CPI (Services ex-Housing ex-Energy)":       ("FRED_YOY", "CUSR0000SASLE"),
    "CPI - Food (YoY)":                                    ("FRED_YOY", "CPIUFDSL"),
    "CPI - Energy (YoY)":                                  ("FRED_YOY", "CPIENGSL"),
    "WTI Crude Oil (Avg)":                                 ("FRED_MONTHLY_AVG", "DCOILWTICO"),
    "Brent Crude Oil (Avg)":                               ("FRED_MONTHLY_AVG", "DCOILBRENTEU"),
    "Copper Price (Avg)":                                  ("FRED_LATEST", "PCOPPUSDM"),
    "PJM Wholesale Electricity (Avg)":                     ("MANUAL", None),
    "Natural Gas Henry Hub (Avg)":                         ("FRED_MONTHLY_AVG", "DHHNGSP"),
    # Consumer purchasing power
    "Real Personal Spending (YoY)":                        ("FRED_YOY", "PCEC96"),
    "Real Avg Hourly Earnings (YoY)":                      ("FRED_YOY_DIFF", ("CES0500000003", "CPIAUCSL")),
    # Labor
    "Initial Jobless Claims (Monthly Avg)":                ("FRED_MONTHLY_AVG", "ICSA"),
    "Continuing Claims (Monthly Avg)":                     ("FRED_MONTHLY_AVG", "CCSA"),
    "JOLTS Job Openings":                                  ("FRED_LATEST", "JTSJOL"),
    "JOLTS Quits Rate":                                    ("FRED_LATEST", "JTSQUR"),
    "JOLTS Layoffs & Discharges":                          ("FRED_LATEST", "JTSLDL"),
    "Challenger Job Cuts (Total)":                         ("MANUAL", None),
    "Challenger Job Cuts (AI/Tech-Related)":               ("MANUAL", None),
    "WARN Act Filings (Major Layoffs)":                    ("MANUAL", None),
    # AI Adoption
    "Anthropic API Price - Frontier (per 1M input tokens)":  ("MANUAL", None),
    "Anthropic API Price - Frontier (per 1M output tokens)": ("MANUAL", None),
    "OpenAI API Price - Frontier (per 1M input tokens)":     ("MANUAL", None),
    "OpenAI API Price - Frontier (per 1M output tokens)":    ("MANUAL", None),
    "Google Gemini API - Frontier (per 1M input tokens)":    ("MANUAL", None),
    "ChatGPT Pro/Plus Subscribers (est.)":                   ("MANUAL", None),
}

QUARTERLY_METRIC_MAP: dict[str, tuple[str, object]] = {
    # Productivity
    "Nonfarm Business Productivity (YoY %)":     ("FRED_QTRLY_YOY", "OPHNFB"),
    "Unit Labor Costs (YoY %)":                  ("FRED_QTRLY_YOY", "ULCNFB"),
    # Disinflationary-offset signal: when ULC growth < productivity growth,
    # this goes negative → productivity is absorbing wage growth → pricing
    # power weakens → core services inflation can decelerate. Powell cites
    # this regularly. Negative = disinflationary; positive = sticky pressure.
    "ULC minus Productivity (YoY %)":            ("FRED_QTRLY_YOY_DIFF", ("ULCNFB", "OPHNFB")),
    "Output Per Hour Worked (Index)":            ("FRED_QTRLY_LATEST", "OPHNFB"),
    "Multifactor Productivity (YoY %)":          ("MANUAL", None),  # BLS publishes annually only
    # Labor (quarterly)
    "Unemployment Rate (U-3)":                   ("FRED_QTRLY_AVG", "UNRATE"),
    "Underemployment Rate (U-6)":                ("FRED_QTRLY_AVG", "U6RATE"),
    "Employment - Information Sector":           ("FRED_QTRLY_AVG", "USINFO"),
    "Employment - Professional & Business Services": ("FRED_QTRLY_AVG", "USPBS"),
    "Employment - Financial Activities":         ("FRED_QTRLY_AVG", "USFIRE"),
    # Hyperscaler Capex
    "Microsoft - Total Capex":                   ("YAHOO_CAPEX_QTRLY", "MSFT"),
    "Microsoft - Capex/Revenue %":               ("YAHOO_CAPEX_PCT", "MSFT"),
    "Alphabet/Google - Total Capex":             ("YAHOO_CAPEX_QTRLY", "GOOG"),
    "Alphabet/Google - Capex/Revenue %":         ("YAHOO_CAPEX_PCT", "GOOG"),
    "Amazon (AWS focus) - Total Capex":          ("YAHOO_CAPEX_QTRLY", "AMZN"),
    "Amazon - Capex/Revenue %":                  ("YAHOO_CAPEX_PCT", "AMZN"),
    "Meta - Total Capex":                        ("YAHOO_CAPEX_QTRLY", "META"),
    "Meta - Capex/Revenue %":                    ("YAHOO_CAPEX_PCT", "META"),
    "Combined Hyperscaler Capex":                ("YAHOO_CAPEX_COMBINED", ["MSFT", "GOOG", "AMZN", "META"]),
    # Semiconductors — segment data not in Yahoo's structured payload
    "NVIDIA Data Center Revenue":                ("MANUAL", None),
    "NVIDIA Data Center Revenue Growth (QoQ %)": ("DERIVED_QOQ", "NVIDIA Data Center Revenue"),
    "AMD Data Center Revenue":                   ("MANUAL", None),
    "TSMC Revenue (Advanced Nodes %)":           ("MANUAL", None),
    # Corporate margins — FactSet paywall
    "S&P 500 Operating Margin (Overall)":        ("MANUAL", None),
    "S&P 500 Operating Margin - Tech Sector":    ("MANUAL", None),
    "S&P 500 Operating Margin - Financials":     ("MANUAL", None),
    "S&P 500 EPS Growth (YoY %)":                ("MANUAL", None),
    # AI company financials — press leaks
    "OpenAI ARR (est.)":                         ("MANUAL", None),
    "OpenAI Gross Margin (est.)":                ("MANUAL", None),
    "Anthropic ARR (est.)":                      ("MANUAL", None),
    "Anthropic Valuation (Last Round)":          ("MANUAL", None),
    "OpenAI Valuation (Last Round)":             ("MANUAL", None),
    # Macro rates
    "10Y Treasury Yield (Avg)":                  ("FRED_QTRLY_AVG", "DGS10"),
    "Real 10Y Rate (TIPS)":                      ("FRED_QTRLY_AVG", "DFII10"),
    # ISM discontinued free FRED feed in 2022
    "ISM Manufacturing PMI":                     ("MANUAL", None),
    "ISM Services PMI":                          ("MANUAL", None),
    "Fed Funds Rate (Upper Bound)":              ("FRED_QTRLY_AVG", "DFEDTARU"),
}


# ─── Period helpers ───

_MONTH_NUM = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
              "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}


def month_to_dates(month_str: str) -> tuple[pd.Timestamp, pd.Timestamp]:
    """'Jan-25' → (2025-01-01, 2025-01-31)."""
    abbr, yr2 = month_str.split("-")
    start = pd.Timestamp(2000 + int(yr2), _MONTH_NUM[abbr], 1)
    return start, start + pd.offsets.MonthEnd(0)


def quarter_to_dates(q_str: str) -> tuple[pd.Timestamp, pd.Timestamp]:
    """'Q1-24' → (2024-01-01, 2024-03-31)."""
    q_part, yr2 = q_str.split("-")
    q_num = int(q_part[1])
    start_month = (q_num - 1) * 3 + 1
    start = pd.Timestamp(2000 + int(yr2), start_month, 1)
    return start, start + pd.offsets.QuarterEnd(0)


# ─── FRED ───

_FRED_CACHE: dict[str, pd.Series] = {}


def fetch_fred(series_id: str) -> pd.Series | None:
    if series_id in _FRED_CACHE:
        return _FRED_CACHE[series_id]
    url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
    try:
        r = requests.get(url, headers=UA, timeout=30)
        r.raise_for_status()
        df = pd.read_csv(io.StringIO(r.text))
        date_col = "DATE" if "DATE" in df.columns else df.columns[0]
        val_col = series_id if series_id in df.columns else df.columns[-1]
        df[date_col] = pd.to_datetime(df[date_col])
        df[val_col] = pd.to_numeric(df[val_col], errors="coerce")
        df = df.dropna(subset=[val_col])
        s = pd.Series(df[val_col].values, index=df[date_col].values, name=series_id)
        s = s.sort_index()
        _FRED_CACHE[series_id] = s
        return s
    except Exception as e:
        print(f"  FRED fetch failed for {series_id}: {e}", file=sys.stderr)
        return None


def fred_value_in_period(series_id: str, p_start: pd.Timestamp, p_end: pd.Timestamp,
                         mode: str) -> float | None:
    """mode ∈ {'avg', 'latest', 'yoy'}."""
    s = fetch_fred(series_id)
    if s is None or s.empty:
        return None
    in_period = s[(s.index >= p_start) & (s.index <= p_end)]
    if in_period.empty:
        return None
    if mode == "avg":
        return round(float(in_period.mean()), 2)
    if mode == "latest":
        return round(float(in_period.iloc[-1]), 2)
    if mode == "yoy":
        current = float(in_period.iloc[-1])
        y_start = p_start - pd.DateOffset(years=1)
        y_end = p_end - pd.DateOffset(years=1)
        prior = s[(s.index >= y_start) & (s.index <= y_end)]
        if prior.empty:
            return None
        prior_val = float(prior.iloc[-1])
        if prior_val == 0:
            return None
        return round((current / prior_val - 1) * 100, 2)
    raise ValueError(f"unknown mode: {mode}")


# ─── Yahoo fundamentals-timeseries (line-item financials) ───
# Switched here from v10/quoteSummary because Yahoo restricted line-item access
# on the v10 endpoint (statements return but capitalExpenditures = None). The
# fundamentals-timeseries endpoint still serves raw quarterly capex + revenue.
# Also switched off yfinance because it's heavily rate-limited; this is a
# direct HTTP call with no auth required.

import time as _time

_YAHOO_TS_CACHE: dict[str, dict[str, list[tuple[pd.Timestamp, float]]]] = {}


def fetch_yahoo_timeseries(ticker: str) -> dict[str, list[tuple[pd.Timestamp, float]]]:
    """
    Pull quarterly capex + revenue history for `ticker`. Returns dict:
        {
          "capex":   [(asOfDate, value), ...],
          "revenue": [(asOfDate, value), ...],
        }
    Values are raw dollars (negative for capex per Yahoo convention).
    """
    if ticker in _YAHOO_TS_CACHE:
        return _YAHOO_TS_CACHE[ticker]
    out = {"capex": [], "revenue": []}
    now = int(_time.time())
    # period1 = 2020-01-01 (1577836800) gives 5+ years of history
    url = (f"https://query1.finance.yahoo.com/ws/fundamentals-timeseries/v1/finance/timeseries/{ticker}"
           f"?type=quarterlyCapitalExpenditure,quarterlyTotalRevenue"
           f"&period1=1577836800&period2={now}")
    try:
        r = requests.get(url, headers=UA, timeout=20)
        if r.status_code != 200:
            print(f"  Yahoo timeseries {ticker}: HTTP {r.status_code}", file=sys.stderr)
            _YAHOO_TS_CACHE[ticker] = out
            return out
        results = r.json().get("timeseries", {}).get("result", [])
        for series in results:
            meta = series.get("meta", {})
            series_types = meta.get("type", [])
            if not series_types:
                continue
            stype = series_types[0]
            bucket = None
            if stype == "quarterlyCapitalExpenditure":
                bucket = out["capex"]
            elif stype == "quarterlyTotalRevenue":
                bucket = out["revenue"]
            if bucket is None:
                continue
            for entry in series.get(stype, []) or []:
                ts = entry.get("asOfDate")
                rv = entry.get("reportedValue", {}) or {}
                val = rv.get("raw") if isinstance(rv, dict) else None
                if ts is None or val is None:
                    continue
                bucket.append((pd.Timestamp(ts), float(val)))
    except Exception as e:
        print(f"  Yahoo timeseries {ticker} failed: {e}", file=sys.stderr)
    _YAHOO_TS_CACHE[ticker] = out
    return out


def yahoo_capex_for_quarter(ticker: str, q_start: pd.Timestamp, q_end: pd.Timestamp) -> float | None:
    """Quarterly capex in $B for the fiscal quarter ending in [q_start, q_end]."""
    ts = fetch_yahoo_timeseries(ticker)
    for (end, val) in ts["capex"]:
        if q_start <= end <= q_end:
            # Yahoo reports capex as negative cash outflow; convert to absolute $B
            return round(abs(val) / 1e9, 2)
    return None


def yahoo_revenue_for_quarter(ticker: str, q_start: pd.Timestamp, q_end: pd.Timestamp) -> float | None:
    ts = fetch_yahoo_timeseries(ticker)
    for (end, val) in ts["revenue"]:
        if q_start <= end <= q_end:
            return round(val / 1e9, 2)
    return None


def yahoo_capex_pct(ticker: str, q_start: pd.Timestamp, q_end: pd.Timestamp) -> float | None:
    capex = yahoo_capex_for_quarter(ticker, q_start, q_end)
    rev = yahoo_revenue_for_quarter(ticker, q_start, q_end)
    if capex is None or rev is None or rev == 0:
        return None
    return round((capex / rev) * 100, 2)


def yahoo_capex_combined(tickers: list[str], q_start: pd.Timestamp,
                         q_end: pd.Timestamp) -> float | None:
    total = 0.0
    found = False
    for t in tickers:
        c = yahoo_capex_for_quarter(t, q_start, q_end)
        if c is not None:
            total += c
            found = True
    return round(total, 2) if found else None


# ─── Manual overrides scaffold ───

def ensure_overrides_scaffold() -> dict:
    if OVERRIDES_PATH.exists():
        with open(OVERRIDES_PATH) as f:
            data = json.load(f)
    else:
        data = {}

    changed = False
    for metric, (strat, _) in MONTHLY_METRIC_MAP.items():
        if strat != "MANUAL":
            continue
        bucket = data.setdefault(metric, {})
        for m in MONTHS:
            if m not in bucket:
                bucket[m] = None
                changed = True
    for metric, (strat, _) in QUARTERLY_METRIC_MAP.items():
        if strat != "MANUAL":
            continue
        bucket = data.setdefault(metric, {})
        for q in QUARTERS:
            if q not in bucket:
                bucket[q] = None
                changed = True

    if changed:
        with open(OVERRIDES_PATH, "w") as f:
            json.dump(data, f, indent=2, sort_keys=True)
        print(f"  Updated {OVERRIDES_PATH.name} scaffold")
    return data


# ─── Workbook ops ───

def find_metric_row(ws, label: str) -> int | None:
    for row in range(4, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == label:
            return row
    return None


def find_period_col(ws, period: str) -> int | None:
    for col in range(4, ws.max_column + 1):
        if ws.cell(row=3, column=col).value == period:
            return col
    return None


def write_cell(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BLUE_INPUT_FONT


# ─── Fill loops ───

def fill_monthly(ws, overrides: dict) -> dict:
    stats = {"filled": 0, "manual_missing": 0, "data_unavailable": 0, "label_not_found": 0}
    for metric, (strat, param) in MONTHLY_METRIC_MAP.items():
        row = find_metric_row(ws, metric)
        if row is None:
            print(f"  WARN: '{metric}' not found in Monthly Metrics sheet")
            stats["label_not_found"] += 1
            continue
        for month in MONTHS:
            col = find_period_col(ws, month)
            if col is None:
                continue
            p_start, p_end = month_to_dates(month)
            value = None
            if strat == "FRED_YOY":
                value = fred_value_in_period(param, p_start, p_end, "yoy")
            elif strat == "FRED_YOY_DIFF":
                s1, s2 = param
                y1 = fred_value_in_period(s1, p_start, p_end, "yoy")
                y2 = fred_value_in_period(s2, p_start, p_end, "yoy")
                value = round(y1 - y2, 2) if (y1 is not None and y2 is not None) else None
            elif strat == "FRED_MONTHLY_AVG":
                value = fred_value_in_period(param, p_start, p_end, "avg")
            elif strat == "FRED_LATEST":
                value = fred_value_in_period(param, p_start, p_end, "latest")
            elif strat == "MANUAL":
                value = overrides.get(metric, {}).get(month)
                if value is None:
                    stats["manual_missing"] += 1
                    continue

            if value is None:
                stats["data_unavailable"] += 1
                continue
            write_cell(ws, row, col, value)
            stats["filled"] += 1
    return stats


def fill_quarterly(ws, overrides: dict) -> dict:
    stats = {"filled": 0, "manual_missing": 0, "data_unavailable": 0, "label_not_found": 0}
    # Pass 1: everything except DERIVED
    for metric, (strat, param) in QUARTERLY_METRIC_MAP.items():
        if strat == "DERIVED_QOQ":
            continue
        row = find_metric_row(ws, metric)
        if row is None:
            print(f"  WARN: '{metric}' not found in Quarterly Metrics sheet")
            stats["label_not_found"] += 1
            continue
        for q in QUARTERS:
            col = find_period_col(ws, q)
            if col is None:
                continue
            p_start, p_end = quarter_to_dates(q)
            value = None
            if strat == "FRED_QTRLY_AVG":
                value = fred_value_in_period(param, p_start, p_end, "avg")
            elif strat == "FRED_QTRLY_YOY":
                value = fred_value_in_period(param, p_start, p_end, "yoy")
            elif strat == "FRED_QTRLY_YOY_DIFF":
                s1, s2 = param
                y1 = fred_value_in_period(s1, p_start, p_end, "yoy")
                y2 = fred_value_in_period(s2, p_start, p_end, "yoy")
                value = round(y1 - y2, 2) if (y1 is not None and y2 is not None) else None
            elif strat == "FRED_QTRLY_LATEST":
                value = fred_value_in_period(param, p_start, p_end, "latest")
            elif strat == "YAHOO_CAPEX_QTRLY":
                value = yahoo_capex_for_quarter(param, p_start, p_end)
            elif strat == "YAHOO_CAPEX_PCT":
                value = yahoo_capex_pct(param, p_start, p_end)
            elif strat == "YAHOO_CAPEX_COMBINED":
                value = yahoo_capex_combined(param, p_start, p_end)
            elif strat == "MANUAL":
                value = overrides.get(metric, {}).get(q)
                if value is None:
                    stats["manual_missing"] += 1
                    continue

            if value is None:
                stats["data_unavailable"] += 1
                continue
            write_cell(ws, row, col, value)
            stats["filled"] += 1

    # Pass 2: derived QoQ metrics — read from source rows just written/overridden
    for metric, (strat, param) in QUARTERLY_METRIC_MAP.items():
        if strat != "DERIVED_QOQ":
            continue
        row = find_metric_row(ws, metric)
        src_row = find_metric_row(ws, param)
        if row is None or src_row is None:
            continue
        prev_val = None
        for q in QUARTERS:
            col = find_period_col(ws, q)
            if col is None:
                continue
            cur_val = ws.cell(row=src_row, column=col).value
            if cur_val is None or not isinstance(cur_val, (int, float)):
                prev_val = None
                continue
            if prev_val is not None and prev_val != 0:
                qoq = round((cur_val / prev_val - 1) * 100, 2)
                write_cell(ws, row, col, qoq)
                stats["filled"] += 1
            prev_val = cur_val

    return stats


# ─── Main ───

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    args = parser.parse_args()
    wb_path = Path(args.workbook)
    if not wb_path.exists():
        print(f"Workbook not found: {wb_path}")
        print("Run `python macro_tracker.py` first to generate it.")
        sys.exit(1)

    print(f"  Loading {wb_path.name}…")
    wb = load_workbook(wb_path)

    print(f"  Scaffolding {OVERRIDES_PATH.name}…")
    overrides = ensure_overrides_scaffold()

    # Guard against missing sheets (workbook structure changed by user)
    sheet_names = wb.sheetnames
    if "Monthly Metrics" not in sheet_names or "Quarterly Metrics" not in sheet_names:
        print(f"  ERROR: expected sheets 'Monthly Metrics' and 'Quarterly Metrics'")
        print(f"  Found sheets: {sheet_names}")
        sys.exit(1)

    print("  Filling Monthly Metrics (FRED + manual overrides)…")
    monthly_stats = fill_monthly(wb["Monthly Metrics"], overrides)
    print(f"    {monthly_stats}")

    print("  Filling Quarterly Metrics (FRED + Yahoo + manual overrides)…")
    quarterly_stats = fill_quarterly(wb["Quarterly Metrics"], overrides)
    print(f"    {quarterly_stats}")

    # Save with file-lock mitigation — if workbook is open in Excel on Windows,
    # openpyxl raises PermissionError. Surface a clear message rather than a stacktrace.
    try:
        wb.save(wb_path)
    except PermissionError:
        print(f"  ERROR: cannot save {wb_path.name} — file appears to be open in Excel.")
        print(f"  Close the workbook in Excel and re-run.")
        sys.exit(1)
    print(f"  Saved {wb_path.name}")

    tf = monthly_stats["filled"] + quarterly_stats["filled"]
    tm = monthly_stats["manual_missing"] + quarterly_stats["manual_missing"]
    td = monthly_stats["data_unavailable"] + quarterly_stats["data_unavailable"]
    print()
    print("  Summary:")
    print(f"    Auto-filled:        {tf}")
    print(f"    Manual-needed:      {tm}  (edit {OVERRIDES_PATH.name})")
    print(f"    Data unavailable:   {td}  (typically future months not yet published)")
    print(f"    FRED series cached: {len(_FRED_CACHE)}")
    print(f"    Yahoo tickers:      {len(_YAHOO_TS_CACHE)}")


if __name__ == "__main__":
    main()

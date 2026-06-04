from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── Color palette ───
DARK_BG = "1B2A4A"
HEADER_BG = "2C3E6B"
SECTION_BG = "3A5BA0"
LIGHT_BG = "E8ECF4"
ALT_ROW = "F5F7FC"
WHITE = "FFFFFF"
GREEN_ACCENT = "27AE60"
RED_ACCENT = "E74C3C"
YELLOW_ACCENT = "F39C12"
BLUE_INPUT = "0000FF"
BLACK = "000000"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
section_font = Font(name="Arial", bold=True, color=WHITE, size=10)
label_font = Font(name="Arial", size=10, color=BLACK)
input_font = Font(name="Arial", size=10, color=BLUE_INPUT)
note_font = Font(name="Arial", size=9, italic=True, color="666666")
title_font = Font(name="Arial", bold=True, color=WHITE, size=14)

header_fill = PatternFill("solid", fgColor=HEADER_BG)
section_fill = PatternFill("solid", fgColor=SECTION_BG)
light_fill = PatternFill("solid", fgColor=LIGHT_BG)
alt_fill = PatternFill("solid", fgColor=ALT_ROW)
dark_fill = PatternFill("solid", fgColor=DARK_BG)
yellow_fill = PatternFill("solid", fgColor="FFF9E6")
green_fill = PatternFill("solid", fgColor="E8F5E9")
red_fill = PatternFill("solid", fgColor="FFEBEE")

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_row(ws, row, cols, fill=header_fill):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border

def style_data_row(ws, row, cols, is_alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = label_font
        cell.border = thin_border
        cell.alignment = center if c > 1 else left_wrap
        if is_alt:
            cell.fill = alt_fill

def style_section_row(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = section_fill

# ════════════════════════════════════════════════════
# SHEET 1: MONTHLY METRICS
# ════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Monthly Metrics"
ws1.sheet_properties.tabColor = "2C3E6B"

months = ["Jan-25","Feb-25","Mar-25","Apr-25","May-25","Jun-25","Jul-25","Aug-25","Sep-25","Oct-25","Nov-25","Dec-25",
          "Jan-26","Feb-26","Mar-26","Apr-26","May-26"]

# Title row
ws1.merge_cells("A1:S1")
ws1.cell(row=1, column=1, value="MACRO THESIS TRACKER — MONTHLY METRICS")
ws1.cell(row=1, column=1).font = title_font
ws1.cell(row=1, column=1).fill = dark_fill
ws1.cell(row=1, column=1).alignment = center
for c in range(1, 20):
    ws1.cell(row=1, column=c).fill = dark_fill

# Headers
headers_m = ["Metric", "Source", "Unit"] + months
ncols_m = len(headers_m)
for i, h in enumerate(headers_m, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, ncols_m)

# Metrics data
metrics_monthly = [
    # Section: Inflation
    ("__SECTION__", "INFLATION & PRICE PRESSURE"),
    ("Core PCE (YoY)", "BEA / FRED", "%"),
    ("Core CPI (YoY)", "BLS", "%"),
    ("Supercore CPI (Services ex-Housing ex-Energy)", "BLS", "%"),
    ("CPI - Food (YoY)", "BLS", "%"),
    ("CPI - Energy (YoY)", "BLS", "%"),
    ("WTI Crude Oil (Avg)", "EIA / FRED", "$/bbl"),
    ("Brent Crude Oil (Avg)", "EIA / FRED", "$/bbl"),
    ("Copper Price (Avg)", "COMEX / FRED", "$/lb"),
    ("PJM Wholesale Electricity (Avg)", "PJM Interconnection", "$/MWh"),
    ("Natural Gas Henry Hub (Avg)", "EIA", "$/MMBtu"),
    # Section: Consumer Purchasing Power
    ("__SECTION__", "CONSUMER PURCHASING POWER"),
    ("Real Personal Spending (YoY)", "BEA / FRED", "%"),
    ("Real Avg Hourly Earnings (YoY)", "BLS / FRED (derived)", "%"),
    # Section: Labor
    ("__SECTION__", "LABOR MARKET"),
    ("Initial Jobless Claims (Monthly Avg)", "DOL / FRED", "000s"),
    ("Continuing Claims (Monthly Avg)", "DOL / FRED", "000s"),
    ("JOLTS Job Openings", "BLS", "000s"),
    ("JOLTS Quits Rate", "BLS", "%"),
    ("JOLTS Layoffs & Discharges", "BLS", "000s"),
    ("Challenger Job Cuts (Total)", "Challenger Gray & Christmas", "000s"),
    ("Challenger Job Cuts (AI/Tech-Related)", "Challenger Gray & Christmas", "000s"),
    ("WARN Act Filings (Major Layoffs)", "State DOL aggregators", "Count"),
    # Section: AI Adoption Signals
    ("__SECTION__", "AI ADOPTION SIGNALS (MONTHLY)"),
    ("Anthropic API Price - Frontier (per 1M input tokens)", "Anthropic pricing page", "$"),
    ("Anthropic API Price - Frontier (per 1M output tokens)", "Anthropic pricing page", "$"),
    ("OpenAI API Price - Frontier (per 1M input tokens)", "OpenAI pricing page", "$"),
    ("OpenAI API Price - Frontier (per 1M output tokens)", "OpenAI pricing page", "$"),
    ("Google Gemini API - Frontier (per 1M input tokens)", "Google AI pricing", "$"),
    ("ChatGPT Pro/Plus Subscribers (est.)", "Press reports / Similarweb", "Millions"),
]

row = 4
for item in metrics_monthly:
    if item[0] == "__SECTION__":
        style_section_row(ws1, row, ncols_m, item[1])
        row += 1
        continue
    ws1.cell(row=row, column=1, value=item[0])
    ws1.cell(row=row, column=2, value=item[1])
    ws1.cell(row=row, column=3, value=item[2])
    style_data_row(ws1, row, ncols_m, is_alt=(row % 2 == 0))
    # Mark data cells as input
    for c in range(4, ncols_m + 1):
        ws1.cell(row=row, column=c).font = input_font
    row += 1

# Column widths
ws1.column_dimensions["A"].width = 48
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 10
for c in range(4, ncols_m + 1):
    ws1.column_dimensions[get_column_letter(c)].width = 11

# Freeze panes
ws1.freeze_panes = "D4"

# ════════════════════════════════════════════════════
# SHEET 2: QUARTERLY METRICS
# ════════════════════════════════════════════════════
ws2 = wb.create_sheet("Quarterly Metrics")
ws2.sheet_properties.tabColor = "27AE60"

quarters = ["Q1-24","Q2-24","Q3-24","Q4-24","Q1-25","Q2-25","Q3-25","Q4-25","Q1-26","Q2-26"]

ws2.merge_cells("A1:M1")
ws2.cell(row=1, column=1, value="MACRO THESIS TRACKER — QUARTERLY METRICS")
ws2.cell(row=1, column=1).font = title_font
ws2.cell(row=1, column=1).fill = dark_fill
ws2.cell(row=1, column=1).alignment = center
for c in range(1, 14):
    ws2.cell(row=1, column=c).fill = dark_fill

headers_q = ["Metric", "Source", "Unit"] + quarters
ncols_q = len(headers_q)
for i, h in enumerate(headers_q, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, ncols_q)

metrics_quarterly = [
    ("__SECTION__", "PRODUCTIVITY & UNIT COSTS"),
    ("Nonfarm Business Productivity (YoY %)", "BLS", "%"),
    ("Unit Labor Costs (YoY %)", "BLS", "%"),
    ("ULC minus Productivity (YoY %)", "BLS (derived)", "%"),
    ("Output Per Hour Worked (Index)", "BLS", "Index"),
    ("Multifactor Productivity (YoY %)", "BLS", "%"),
    
    ("__SECTION__", "LABOR MARKET (QUARTERLY)"),
    ("Unemployment Rate (U-3)", "BLS", "%"),
    ("Underemployment Rate (U-6)", "BLS", "%"),
    ("Employment - Information Sector", "BLS", "000s"),
    ("Employment - Professional & Business Services", "BLS", "000s"),
    ("Employment - Financial Activities", "BLS", "000s"),
    
    ("__SECTION__", "HYPERSCALER CAPEX ($B)"),
    ("Microsoft - Total Capex", "MSFT 10-Q/K", "$B"),
    ("Microsoft - Capex/Revenue %", "MSFT 10-Q/K", "%"),
    ("Alphabet/Google - Total Capex", "GOOG 10-Q/K", "$B"),
    ("Alphabet/Google - Capex/Revenue %", "GOOG 10-Q/K", "%"),
    ("Amazon (AWS focus) - Total Capex", "AMZN 10-Q/K", "$B"),
    ("Amazon - Capex/Revenue %", "AMZN 10-Q/K", "%"),
    ("Meta - Total Capex", "META 10-Q/K", "$B"),
    ("Meta - Capex/Revenue %", "META 10-Q/K", "%"),
    ("Combined Hyperscaler Capex", "Calculated", "$B"),
    
    ("__SECTION__", "SEMICONDUCTOR / COMPUTE DEMAND"),
    ("NVIDIA Data Center Revenue", "NVDA 10-Q/K", "$B"),
    ("NVIDIA Data Center Revenue Growth (QoQ %)", "NVDA 10-Q/K", "%"),
    ("AMD Data Center Revenue", "AMD 10-Q/K", "$B"),
    ("TSMC Revenue (Advanced Nodes %)", "TSMC Monthly", "%"),
    
    ("__SECTION__", "CORPORATE MARGINS & PROFITABILITY"),
    ("S&P 500 Operating Margin (Overall)", "FactSet / Bloomberg", "%"),
    ("S&P 500 Operating Margin - Tech Sector", "FactSet / Bloomberg", "%"),
    ("S&P 500 Operating Margin - Financials", "FactSet / Bloomberg", "%"),
    ("S&P 500 EPS Growth (YoY %)", "FactSet / Bloomberg", "%"),
    
    ("__SECTION__", "AI COMPANY FINANCIALS (PRIVATE - EST.)"),
    ("OpenAI ARR (est.)", "Press / Leaks", "$B"),
    ("OpenAI Gross Margin (est.)", "Press / Leaks", "%"),
    ("Anthropic ARR (est.)", "Press / Leaks", "$B"),
    ("Anthropic Valuation (Last Round)", "Crunchbase / PitchBook", "$B"),
    ("OpenAI Valuation (Last Round)", "Crunchbase / PitchBook", "$B"),
    
    ("__SECTION__", "MACRO RATES & CONDITIONS"),
    ("10Y Treasury Yield (Avg)", "FRED", "%"),
    ("Real 10Y Rate (TIPS)", "FRED", "%"),
    ("ISM Manufacturing PMI", "ISM", "Index"),
    ("ISM Services PMI", "ISM", "Index"),
    ("Fed Funds Rate (Upper Bound)", "FOMC", "%"),
]

row = 4
for item in metrics_quarterly:
    if item[0] == "__SECTION__":
        style_section_row(ws2, row, ncols_q, item[1])
        row += 1
        continue
    ws2.cell(row=row, column=1, value=item[0])
    ws2.cell(row=row, column=2, value=item[1])
    ws2.cell(row=row, column=3, value=item[2])
    style_data_row(ws2, row, ncols_q, is_alt=(row % 2 == 0))
    for c in range(4, ncols_q + 1):
        ws2.cell(row=row, column=c).font = input_font
    row += 1

ws2.column_dimensions["A"].width = 48
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 10
for c in range(4, ncols_q + 1):
    ws2.column_dimensions[get_column_letter(c)].width = 12
ws2.freeze_panes = "D4"

# ════════════════════════════════════════════════════
# SHEET 3: AI COST CROSSOVER ANALYSIS
# ════════════════════════════════════════════════════
ws3 = wb.create_sheet("AI Cost Crossover")
ws3.sheet_properties.tabColor = "F39C12"

ws3.merge_cells("A1:H1")
ws3.cell(row=1, column=1, value="AI vs. HUMAN COST CROSSOVER ANALYSIS")
ws3.cell(row=1, column=1).font = title_font
ws3.cell(row=1, column=1).fill = dark_fill
ws3.cell(row=1, column=1).alignment = center
for c in range(1, 9):
    ws3.cell(row=1, column=c).fill = dark_fill

# Assumptions block
style_section_row(ws3, 3, 8, "ASSUMPTIONS (Blue = Editable)")
assumptions = [
    ("Avg tasks per employee per day", 40, "B4"),
    ("Working days per year", 250, "B5"),
    ("Avg tokens per AI task (input)", 2000, "B6"),
    ("Avg tokens per AI task (output)", 500, "B7"),
]
for i, (label, val, _) in enumerate(assumptions):
    r = 4 + i
    ws3.cell(row=r, column=1, value=label).font = label_font
    ws3.cell(row=r, column=2, value=val).font = input_font
    ws3.cell(row=r, column=1).fill = yellow_fill
    ws3.cell(row=r, column=2).fill = yellow_fill
    for c in range(1, 9):
        ws3.cell(row=r, column=c).border = thin_border

# Job roles table
style_section_row(ws3, 9, 8, "COST COMPARISON BY ROLE")
role_headers = ["Job Role", "Fully Loaded Annual Cost ($)", "Tasks/Day", "AI Cost/Task ($)", "Annual AI Cost ($)", "Savings ($)", "Savings %", "Breakeven API Price ($/1M tokens)"]
for i, h in enumerate(role_headers, 1):
    ws3.cell(row=10, column=i, value=h)
style_header_row(ws3, 10, 8)

roles = [
    ("Customer Service Agent", 55000, 50),
    ("Data Entry Clerk", 42000, 60),
    ("Junior Analyst", 75000, 30),
    ("Content Writer", 65000, 20),
    ("Legal Paralegal", 70000, 25),
    ("IT Help Desk (L1)", 52000, 45),
    ("Bookkeeper", 50000, 35),
    ("Insurance Claims Processor", 58000, 40),
]

for i, (role, cost, tasks) in enumerate(roles):
    r = 11 + i
    ws3.cell(row=r, column=1, value=role)
    ws3.cell(row=r, column=2, value=cost).font = input_font
    ws3.cell(row=r, column=2).number_format = '$#,##0'
    ws3.cell(row=r, column=3, value=tasks).font = input_font
    # AI cost per task = (input_tokens/1M * input_price + output_tokens/1M * output_price)
    # Using placeholder formula referencing a price input
    ws3.cell(row=r, column=4, value=f'=(B6/1000000*15 + B7/1000000*75)')  # ~$0.0675 per task at current frontier
    ws3.cell(row=r, column=4).number_format = '$#,##0.000'
    ws3.cell(row=r, column=5, value=f'=D{r}*C{r}*B5')  # Annual AI cost
    ws3.cell(row=r, column=5).number_format = '$#,##0'
    ws3.cell(row=r, column=6, value=f'=B{r}-E{r}')  # Savings
    ws3.cell(row=r, column=6).number_format = '$#,##0'
    ws3.cell(row=r, column=7, value=f'=IF(B{r}=0,0,F{r}/B{r})')  # Savings %
    ws3.cell(row=r, column=7).number_format = '0.0%'
    ws3.cell(row=r, column=8, value=f'=IF(C{r}*B5=0,0, B{r}/(C{r}*B5*(B6+B7))*1000000)')  # Breakeven
    ws3.cell(row=r, column=8).number_format = '$#,##0.00'
    style_data_row(ws3, r, 8, is_alt=(i % 2 == 0))

for c in [1]:
    ws3.column_dimensions[get_column_letter(c)].width = 35
for c in range(2, 9):
    ws3.column_dimensions[get_column_letter(c)].width = 22
ws3.freeze_panes = "B11"

# ════════════════════════════════════════════════════
# SHEET 4: THESIS DASHBOARD
# ════════════════════════════════════════════════════
ws4 = wb.create_sheet("Thesis Dashboard")
ws4.sheet_properties.tabColor = "E74C3C"

ws4.merge_cells("A1:F1")
ws4.cell(row=1, column=1, value="THESIS STATUS DASHBOARD")
ws4.cell(row=1, column=1).font = title_font
ws4.cell(row=1, column=1).fill = dark_fill
ws4.cell(row=1, column=1).alignment = center
for c in range(1, 7):
    ws4.cell(row=1, column=c).fill = dark_fill

# Thesis legs
style_section_row(ws4, 3, 6, "THESIS LEGS — STATUS TRACKING")
dash_headers = ["Thesis Leg", "Current Status", "Direction (▲▼▬)", "Key Metric to Watch", "Falsification Signal", "Notes / Latest Update"]
for i, h in enumerate(dash_headers, 1):
    ws4.cell(row=4, column=i, value=h)
style_header_row(ws4, 4, 6)

thesis_legs = [
    ("1. Oil stays elevated ($85-90)", "Enter status", "▬", "WTI / Brent price", "Oil drops below $70 sustained for 3+ months", ""),
    ("2. Core inflation stays sticky", "Enter status", "▬", "Supercore CPI, Core PCE", "Supercore CPI drops below 3% YoY for 2+ months", ""),
    ("3. Productivity boom NOT yet here", "Enter status", "▬", "Nonfarm Productivity, Unit Labor Costs", "Productivity growth >3% for 2+ quarters w/ falling ULC", ""),
    ("4. Big tech leads AI layoffs", "Enter status", "▬", "Challenger cuts (AI-tagged), JOLTS info sector", "No increase in AI-related layoff announcements by mid-2026", ""),
    ("5. Small/mid cos wait on AI adoption", "Enter status", "▬", "JOLTS openings (small biz), NFIB hiring plans", "Small biz AI adoption >40% in surveys w/ active workforce cuts", ""),
    ("6. AI capex continues rising", "Enter status", "▬", "Hyperscaler capex, NVDA DC revenue", "Combined hyperscaler capex declines QoQ for 2+ quarters", ""),
    ("7. AI pricing must keep falling", "Enter status", "▬", "API $/1M tokens (frontier models)", "API prices plateau or rise for 6+ months", ""),
    ("8. AI cos need public markets", "Enter status", "▬", "IPO filings, funding rounds, burn rate reports", "Major AI cos achieve profitability pre-IPO", ""),
    ("9. Electricity becomes bottleneck", "Enter status", "▬", "PJM prices, utility interconnection queue", "Wholesale electricity falls below 2024 levels near DC hubs", ""),
    ("10. Eventual disinflation wave (2-4 yr)", "Enter status", "▬", "Productivity + ULC trend, AI penetration rate", "Inflation accelerates despite rising AI adoption", ""),
]

for i, (leg, status, direction, metric, falsification, notes) in enumerate(thesis_legs):
    r = 5 + i
    ws4.cell(row=r, column=1, value=leg)
    ws4.cell(row=r, column=2, value=status).font = input_font
    ws4.cell(row=r, column=3, value=direction).font = input_font
    ws4.cell(row=r, column=4, value=metric)
    ws4.cell(row=r, column=5, value=falsification)
    ws4.cell(row=r, column=6, value=notes).font = input_font
    style_data_row(ws4, r, 6, is_alt=(i % 2 == 0))
    ws4.cell(row=r, column=2).fill = yellow_fill
    ws4.cell(row=r, column=3).fill = yellow_fill
    ws4.cell(row=r, column=6).fill = yellow_fill

# Key ratios section
key_row = 5 + len(thesis_legs) + 2
style_section_row(ws4, key_row, 6, "KEY RATIOS & COMPUTED SIGNALS")
ratio_headers = ["Ratio / Signal", "Formula", "Current Value", "Threshold", "Status", "Interpretation"]
for i, h in enumerate(ratio_headers, 1):
    ws4.cell(row=key_row + 1, column=i, value=h)
style_header_row(ws4, key_row + 1, 6)

ratios = [
    ("Hyperscaler Capex / Revenue", "Combined capex / combined revenue", "", ">15% = aggressive buildout", "", "Rising = still in buildout phase"),
    ("NVDA DC Rev Growth (QoQ)", "Current / Prior - 1", "", "<10% QoQ = demand plateau", "", "Decelerating = compute demand peaking"),
    ("AI API Price Index (Normalized)", "Current frontier price / Jan-24 price", "", "<0.3 = adoption-enabling", "", "Falling fast = thesis on track"),
    ("Job Openings / Unemployment", "JOLTS openings / unemployed persons", "", "<1.0 = labor market loosening", "", "Falling = slack building"),
    ("Supercore CPI Trend (3m avg)", "3-month moving average", "", "<3.0% = disinflation starting", "", "Persistent >3.5% = sticky inflation"),
    ("AI Layoff Intensity", "AI-tagged cuts / total cuts", "", ">20% = AI displacement wave", "", "Rising share = thesis accelerating"),
    ("Unit Labor Cost vs Productivity", "ULC growth - Productivity growth", "", "<0 = productivity winning", "", "Negative = deflationary pressure"),
    ("Electricity Cost Pressure", "PJM YoY % change", "", ">20% = bottleneck forming", "", "Rising = AI infra cost headwind"),
]

for i, (name, formula, val, threshold, status, interp) in enumerate(ratios):
    r = key_row + 2 + i
    ws4.cell(row=r, column=1, value=name)
    ws4.cell(row=r, column=2, value=formula)
    ws4.cell(row=r, column=3, value=val).font = input_font
    ws4.cell(row=r, column=4, value=threshold)
    ws4.cell(row=r, column=5, value=status).font = input_font
    ws4.cell(row=r, column=6, value=interp)
    style_data_row(ws4, r, 6, is_alt=(i % 2 == 0))
    ws4.cell(row=r, column=3).fill = yellow_fill
    ws4.cell(row=r, column=5).fill = yellow_fill

ws4.column_dimensions["A"].width = 38
ws4.column_dimensions["B"].width = 35
ws4.column_dimensions["C"].width = 16
ws4.column_dimensions["D"].width = 32
ws4.column_dimensions["E"].width = 16
ws4.column_dimensions["F"].width = 42
ws4.freeze_panes = "B5"

# ════════════════════════════════════════════════════
# SHEET 5: DATA SOURCES REFERENCE
# ════════════════════════════════════════════════════
ws5 = wb.create_sheet("Data Sources")
ws5.sheet_properties.tabColor = "8E44AD"

ws5.merge_cells("A1:E1")
ws5.cell(row=1, column=1, value="DATA SOURCES & ACCESS REFERENCE")
ws5.cell(row=1, column=1).font = title_font
ws5.cell(row=1, column=1).fill = dark_fill
ws5.cell(row=1, column=1).alignment = center
for c in range(1, 6):
    ws5.cell(row=1, column=c).fill = dark_fill

src_headers = ["Metric Category", "Source", "URL / Access", "Frequency", "Notes"]
for i, h in enumerate(src_headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, 5)

sources = [
    ("Core PCE", "BEA via FRED", "https://fred.stlouisfed.org/series/PCEPILFE", "Monthly (last Friday of month)", "Released ~30 days after month end"),
    ("Core CPI / Supercore", "BLS", "https://www.bls.gov/cpi/", "Monthly (2nd week)", "Supercore = CPI Services ex-Shelter ex-Energy"),
    ("Real Personal Spending", "BEA via FRED", "https://fred.stlouisfed.org/series/PCEC96", "Monthly (last business day)", "Real PCE in chained 2017 dollars; YoY = real consumption growth"),
    ("Real Avg Hourly Earnings", "BLS via FRED (derived)", "CES0500000003 (nom AHE) - CPIAUCSL", "Monthly (1st Friday)", "Derived as nom AHE YoY minus CPI YoY"),
    ("WTI / Brent Oil", "EIA via FRED", "https://fred.stlouisfed.org/series/DCOILWTICO", "Daily", "Use monthly average for tracker"),
    ("Copper", "COMEX via FRED", "https://fred.stlouisfed.org/series/PCOPPUSDM", "Monthly", "Global copper price per pound"),
    ("PJM Electricity", "PJM Interconnection", "https://dataminer2.pjm.com/", "Daily/Hourly", "Use DA LMP for major hubs"),
    ("Initial/Continuing Claims", "DOL via FRED", "https://fred.stlouisfed.org/series/ICSA", "Weekly (Thursday)", "Avg weekly for monthly figure"),
    ("JOLTS", "BLS", "https://www.bls.gov/jlt/", "Monthly (~2 month lag)", "Openings, quits, layoffs by sector"),
    ("Challenger Job Cuts", "Challenger Gray & Christmas", "https://www.challengergray.com/", "Monthly (1st Thursday)", "Filter by reason for AI-related"),
    ("Nonfarm Productivity", "BLS", "https://www.bls.gov/lpc/", "Quarterly (~35 day lag)", "Preliminary, then revised"),
    ("Unit Labor Costs", "BLS", "https://www.bls.gov/lpc/", "Quarterly", "Released with productivity"),
    ("Hyperscaler Capex", "SEC Filings (10-Q/K)", "https://www.sec.gov/cgi-bin/browse-edgar", "Quarterly", "MSFT, GOOG, AMZN, META earnings"),
    ("NVIDIA DC Revenue", "NVDA Earnings", "https://investor.nvidia.com/", "Quarterly", "Data Center segment in earnings release"),
    ("API Pricing - Anthropic", "Anthropic", "https://docs.anthropic.com/en/docs/about-claude/models", "As updated", "Track flagship model pricing"),
    ("API Pricing - OpenAI", "OpenAI", "https://openai.com/api/pricing/", "As updated", "Track GPT-4o / frontier pricing"),
    ("Unemployment (U-3, U-6)", "BLS", "https://www.bls.gov/cps/", "Monthly (1st Friday)", "U-6 includes discouraged + part-time"),
    ("ISM PMI", "ISM", "https://www.ismworld.org/", "Monthly (1st business day)", "Manufacturing and Services separate"),
    ("10Y Treasury / TIPS", "FRED", "https://fred.stlouisfed.org/series/DGS10", "Daily", "Real rate = nominal - breakeven"),
    ("S&P 500 Margins", "FactSet Earnings Insight", "https://www.factset.com/earningsinsight", "Weekly", "Free weekly PDF report"),
    ("WARN Act Filings", "State DOL websites", "Varies by state", "As filed", "60-day advance notice for large layoffs"),
    ("AI Company Valuations", "Crunchbase / PitchBook", "https://www.crunchbase.com/", "As rounds close", "Private company data, may require subscription"),
]

for i, (cat, src, url, freq, notes) in enumerate(sources):
    r = 4 + i
    ws5.cell(row=r, column=1, value=cat)
    ws5.cell(row=r, column=2, value=src)
    ws5.cell(row=r, column=3, value=url)
    ws5.cell(row=r, column=4, value=freq)
    ws5.cell(row=r, column=5, value=notes)
    style_data_row(ws5, r, 5, is_alt=(i % 2 == 0))

ws5.column_dimensions["A"].width = 30
ws5.column_dimensions["B"].width = 28
ws5.column_dimensions["C"].width = 55
ws5.column_dimensions["D"].width = 28
ws5.column_dimensions["E"].width = 45
ws5.freeze_panes = "A4"

# Save
wb.save("macro_thesis_tracker.xlsx")
print("Done — macro_thesis_tracker.xlsx created")
